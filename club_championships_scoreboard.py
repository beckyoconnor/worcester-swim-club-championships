#!/usr/bin/env python3
"""
Club Championships Scoreboard
Calculates age group trophies based on championship rules.

Rules:
- Top 8 events across at least 5 of the 6 categories
- Under 12s: max 3 races per category
- 12 and over: max 2 races per category
- Highest collated FINA/WA points wins
"""

import pandas as pd
import os
from typing import Dict, List, Tuple
from openpyxl import load_workbook
import glob


def get_event_gender_map(excel_file: str) -> Dict[str, str]:
    """
    Map event numbers to gender by reading event names from Excel.
    
    Args:
        excel_file: Path to Excel file
        
    Returns:
        Dictionary mapping event number to gender ('Male' or 'Female')
    """
    wb = load_workbook(excel_file, data_only=True)
    numbered_sheets = [sheet for sheet in wb.sheetnames if sheet.isdigit()]
    
    event_gender_map = {}
    
    for sheet_name in numbered_sheets:
        ws = wb[sheet_name]
        event_name = ws.cell(1, 1).value
        event_name_row2 = ws.cell(2, 1).value
        
        combined_text = ''
        if event_name:
            combined_text += str(event_name).lower() + ' '
        if event_name_row2:
            combined_text += str(event_name_row2).lower()
        
        if combined_text:
            if 'female' in combined_text or 'girl' in combined_text:
                event_gender_map[sheet_name] = 'Female'
            elif 'male' in combined_text or 'open/male' in combined_text or 'boy' in combined_text:
                event_gender_map[sheet_name] = 'Male'
            elif 'open' in combined_text:
                event_gender_map[sheet_name] = 'Male'
            else:
                event_gender_map[sheet_name] = 'Unknown'
        else:
            event_gender_map[sheet_name] = 'Unknown'
    
    wb.close()
    return event_gender_map


def get_event_gender_map_from_csvs(folder: str) -> Dict[str, str]:
    """Fallback: infer event gender by reading event CSV filenames and header text.

    Looks under cleaned_files/ if present, otherwise the folder itself.
    """
    cleaned_folder = os.path.join(folder, 'cleaned_files')
    search_folder = cleaned_folder if os.path.exists(cleaned_folder) else folder

    csv_files = glob.glob(os.path.join(search_folder, 'event_*.csv'))
    event_gender_map: Dict[str, str] = {}
    for csv_file in csv_files:
        try:
            basename = os.path.basename(csv_file)
            event_number = os.path.splitext(basename)[0].split('_')[-1]
            df = pd.read_csv(csv_file)
            if 'Event Name' in df.columns and len(df) > 0:
                event_name = str(df['Event Name'].iloc[0]).lower()
                if 'female' in event_name or 'girl' in event_name:
                    event_gender_map[event_number] = 'Female'
                elif 'male' in event_name or 'open/male' in event_name or 'boy' in event_name:
                    event_gender_map[event_number] = 'Male'
                elif 'open' in event_name:
                    event_gender_map[event_number] = 'Male'
                else:
                    event_gender_map[event_number] = 'Male'
            else:
                event_gender_map[event_number] = 'Unknown'
        except Exception:
            event_gender_map[event_number] = 'Unknown'
    return event_gender_map

def load_all_events(folder: str) -> pd.DataFrame:
    """
    Load all event CSV files into a single dataframe.
    
    Args:
        folder: Path to folder containing event CSV files
        
    Returns:
        Combined dataframe with all events
    """
    # Look for event files in cleaned_files subfolder
    cleaned_folder = os.path.join(folder, 'cleaned_files')
    if os.path.exists(cleaned_folder):
        search_folder = cleaned_folder
    else:
        search_folder = folder
    
    csv_files = [f for f in os.listdir(search_folder) if f.startswith('event_') and f.endswith('.csv')]
    
    def _norm_time_str(val: str) -> str:
        """Normalize time to HH:MM:SS.HS (hours:minutes:seconds.hundredths).

        Accepted inputs:
        - SS.HS (e.g., '31.95')
        - MM:SS.HS (e.g., '1:57.04') or MM:SS (e.g., '4:17')
        - HH:MM:SS.HS or HH:MM:SS
        - Numeric seconds (e.g., '31')
        Returns '-' for blanks or unparsable values.
        """
        import re
        try:
            if pd.isna(val):
                return '-'
            s = str(val).strip()
            if not s or s == '-':
                return '-'

            # HH:MM:SS(.H+)? - accept any decimals, trim to 2
            m = re.match(r"^(\d{1,2}):(\d{1,2}):(\d{1,2})(?:\.(\d+))?$", s)
            if m:
                hh = int(m.group(1))
                mm = int(m.group(2))
                ss = int(m.group(3))
                dec = m.group(4) or "0"
                hs = (dec[:2]).ljust(2, '0')
                return f"{hh:02d}:{mm:02d}:{ss:02d}.{hs}"

            # MM:SS(.H+)?
            m = re.match(r"^(\d{1,2}):(\d{1,2})(?:\.(\d+))?$", s)
            if m:
                mm = int(m.group(1))
                ss = int(m.group(2))
                dec = m.group(3) or "0"
                hs = (dec[:2]).ljust(2, '0')
                return f"00:{mm:02d}:{ss:02d}.{hs}"

            # SS.H+ (any decimals)
            m = re.match(r"^(\d{1,2})\.(\d+)$", s)
            if m:
                ss = int(m.group(1))
                dec = m.group(2)
                hs = (dec[:2]).ljust(2, '0')
                return f"00:00:{ss:02d}.{hs}"

            # MM:SS with single-digit seconds etc. already covered; if only digits or digits.decimals treat as seconds
            if re.match(r"^\d+$", s):
                ss = int(s)
                return f"00:00:{ss:02d}.00"
            if re.match(r"^\d+\.\d+$", s):
                parts = s.split('.')
                ss = int(parts[0])
                hs = int(parts[1])
                return f"00:00:{ss:02d}.{hs:02d}"

            # Already normalized
            if re.match(r"^\d{2}:\d{2}:\d{2}\.\d{2}$", s):
                return s

            return '-'
        except Exception:
            return '-'

    dfs = []
    for csv_file in csv_files:
        file_path = os.path.join(search_folder, csv_file)
        df = pd.read_csv(file_path)
        if 'Time' in df.columns:
            df['Time'] = df['Time'].apply(_norm_time_str)
        dfs.append(df)
    
    return pd.concat(dfs, ignore_index=True)


def export_all_events_union(base_folder: str, df_all: pd.DataFrame) -> None:
    """Write a single unioned events file to championship_results/.

    Tries Parquet first (fastest, smallest); falls back to CSV if Parquet engine
    is unavailable.
    """
    try:
        out_dir = os.path.join(base_folder, 'championship_results')
        os.makedirs(out_dir, exist_ok=True)

        # Optimize dtypes prior to write
        df = df_all.copy()
        if 'Event Number' in df.columns:
            df['Event Number'] = df['Event Number'].astype(str)
            df['Event Number'] = df['Event Number'].astype('category')
        for col in ['Event Name', 'Event Category', 'Club']:
            if col in df.columns:
                df[col] = df[col].astype('category')
        if 'Age' in df.columns:
            df['Age'] = pd.to_numeric(df['Age'], errors='coerce').fillna(0).astype('int16')
        if 'WA Points' in df.columns:
            df['WA Points'] = pd.to_numeric(df['WA Points'], errors='coerce').fillna(0).astype('int32')

        parquet_path = os.path.join(out_dir, 'events_all.parquet')
        try:
            df.to_parquet(parquet_path, index=False)
            print(f"✓ Saved unioned events Parquet: {parquet_path} ({len(df)} rows)")
            return
        except Exception as e:
            print(f"⚠️ Could not write Parquet ({e}); falling back to CSV…")
        # Fallback to CSV
        csv_path = os.path.join(out_dir, 'events_all.csv')
        df.to_csv(csv_path, index=False)
        print(f"✓ Saved unioned events CSV: {csv_path} ({len(df)} rows)")
    except Exception as e:
        print(f"⚠️ Failed to export unioned events: {e}")

def calculate_championship_scores(df_all: pd.DataFrame, event_gender_map: Dict[str, str]) -> pd.DataFrame:
    """
    Calculate championship scores based on the rules:
    - Top 8 events across at least 5 categories
    - Under 12s: max 3 races per category
    - 12 and over: max 2 races per category
    
    Args:
        df_all: Dataframe with all events
        event_gender_map: Mapping of event numbers to gender
        
    Returns:
        Dataframe with championship scores
    """
    # Add gender column
    df_all['Event Number'] = df_all['Event Number'].astype(str)
    df_all['Gender'] = df_all['Event Number'].map(event_gender_map)
    
    # Remove Unknown gender entries
    df_all = df_all[df_all['Gender'] != 'Unknown'].copy()
    
    championship_results = []
    
    # Group by swimmer name
    for name, swimmer_events in df_all.groupby('Name'):
        # Get swimmer info
        age = swimmer_events['Age'].iloc[0]
        club = swimmer_events['Club'].iloc[0]
        gender = swimmer_events['Gender'].mode()[0] if len(swimmer_events['Gender'].mode()) > 0 else swimmer_events['Gender'].iloc[0]
        
        # Determine max races per category based on age
        max_per_category = 3 if age < 12 else 2
        
        # Sort by WA Points descending
        swimmer_events_sorted = swimmer_events.sort_values('WA Points', ascending=False)
        
        # For each category, take top N races
        category_events = {}
        for category in ['Sprint', 'Free', '100 Form', '200 Form', 'IM', 'Distance']:
            cat_events = swimmer_events_sorted[swimmer_events_sorted['Event Category'] == category]
            # Remove duplicates - keep only best score per unique event
            cat_events = cat_events.drop_duplicates(subset=['Event Number'], keep='first')
            # Take top N for this category
            top_n = cat_events.head(max_per_category)
            if len(top_n) > 0:
                category_events[category] = top_n
        
        # Check if swimmer competed in at least 5 categories
        num_categories = len(category_events)
        
        if num_categories < 5:
            # Not eligible - didn't compete in at least 5 categories
            continue
        
        # Combine all category events and take top 8
        all_category_events = pd.concat(category_events.values(), ignore_index=True)
        top_8_events = all_category_events.nlargest(8, 'WA Points')
        
        # Calculate totals
        total_points = top_8_events['WA Points'].sum()
        avg_points = top_8_events['WA Points'].mean()
        best_event_points = top_8_events['WA Points'].max()
        num_events = len(top_8_events)
        
        # Count events per category
        category_counts = top_8_events['Event Category'].value_counts().to_dict()
        
        championship_results.append({
            'Name': name,
            'Age': age,
            'Gender': gender,
            'Club': club,
            'Total_Points': total_points,
            'Average_Points': avg_points,
            'Best_Event_Points': best_event_points,
            'Events_Count': num_events,
            'Categories_Competed': num_categories,
            'Sprint_Events': category_counts.get('Sprint', 0),
            'Free_Events': category_counts.get('Free', 0),
            'Form_100_Events': category_counts.get('100 Form', 0),
            'Form_200_Events': category_counts.get('200 Form', 0),
            'IM_Events': category_counts.get('IM', 0),
            'Distance_Events': category_counts.get('Distance', 0)
        })
    
    return pd.DataFrame(championship_results)


def create_age_groups(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create age group categories for the championship.
    
    Args:
        df: Dataframe with championship scores
        
    Returns:
        Dataframe with age group column added
    """
    df = df.copy()
    
    # Define age groups
    bins = [0, 10, 12, 14, 16, 100]
    labels = ['9-10', '11-12', '13-14', '15-16', '17+']
    
    df['Age Group'] = pd.cut(df['Age'], bins=bins, labels=labels, right=True)
    
    return df


def display_scoreboard(df_champs: pd.DataFrame, gender: str, title: str):
    """
    Display championship scoreboard for a specific gender.
    
    Args:
        df_champs: Championship results dataframe
        gender: 'Male' or 'Female'
        title: Title for the scoreboard
    """
    df_gender = df_champs[df_champs['Gender'] == gender].copy()
    
    if len(df_gender) == 0:
        print(f"\nNo eligible swimmers found for {gender}")
        return
    
    # Sort by age group and total points
    df_gender = df_gender.sort_values(['Age Group', 'Total_Points'], ascending=[True, False])
    
    print(f"\n{'='*100}")
    print(f"{title}")
    print('='*100)
    
    current_age_group = None
    
    for _, swimmer in df_gender.iterrows():
        if swimmer['Age Group'] != current_age_group:
            current_age_group = swimmer['Age Group']
            print(f"\n🏆 {current_age_group} AGE GROUP")
            print('-'*100)
            print(f"{'Pos':<4} {'Name':<25} {'Age':<4} {'Club':<15} {'Total':<7} {'Avg':<6} {'Events':<7} {'Categories'}")
            print('-'*100)
            position = 1
        
        print(f"{position:<4} {swimmer['Name']:<25} {swimmer['Age']:<4} {swimmer['Club']:<15} "
              f"{swimmer['Total_Points']:<7.0f} {swimmer['Average_Points']:<6.1f} {swimmer['Events_Count']:<7} "
              f"Sp:{swimmer['Sprint_Events']} Fr:{swimmer['Free_Events']} "
              f"100F:{swimmer['Form_100_Events']} 200F:{swimmer['Form_200_Events']} "
              f"IM:{swimmer['IM_Events']} Dist:{swimmer['Distance_Events']}")
        position += 1


def export_scoreboard(df_champs: pd.DataFrame, output_folder: str):
    """
    Export championship scoreboards to CSV files.
    
    Args:
        df_champs: Championship results dataframe
        output_folder: Folder to save results
    """
    # Create championship_results subfolder
    results_folder = os.path.join(output_folder, 'championship_results')
    os.makedirs(results_folder, exist_ok=True)
    
    # Export boys results
    df_boys = df_champs[df_champs['Gender'] == 'Male'].sort_values(['Age Group', 'Total_Points'], 
                                                                     ascending=[True, False])
    output_file = os.path.join(results_folder, 'championship_scoreboard_boys.csv')
    df_boys_export = df_boys[['Age Group', 'Name', 'Age', 'Club', 'Total_Points', 'Average_Points',
                               'Best_Event_Points', 'Events_Count', 'Categories_Competed',
                               'Sprint_Events', 'Free_Events', 'Form_100_Events', 'Form_200_Events',
                               'IM_Events', 'Distance_Events']]
    df_boys_export.to_csv(output_file, index=False)
    print(f"\n✓ Saved: {output_file} ({len(df_boys)} boys)")
    
    # Export girls results
    df_girls = df_champs[df_champs['Gender'] == 'Female'].sort_values(['Age Group', 'Total_Points'], 
                                                                        ascending=[True, False])
    output_file = os.path.join(results_folder, 'championship_scoreboard_girls.csv')
    df_girls_export = df_girls[['Age Group', 'Name', 'Age', 'Club', 'Total_Points', 'Average_Points',
                                 'Best_Event_Points', 'Events_Count', 'Categories_Competed',
                                 'Sprint_Events', 'Free_Events', 'Form_100_Events', 'Form_200_Events',
                                 'IM_Events', 'Distance_Events']]
    df_girls_export.to_csv(output_file, index=False)
    print(f"✓ Saved: {output_file} ({len(df_girls)} girls)")
    
    # Export age group winners
    winners = []
    for age_group in ['9-10', '11-12', '13-14', '15-16', '17+']:
        for gender in ['Male', 'Female']:
            ag_gender = df_champs[(df_champs['Age Group'] == age_group) & 
                                  (df_champs['Gender'] == gender)]
            if len(ag_gender) > 0:
                winner = ag_gender.nlargest(1, 'Total_Points').iloc[0]
                winners.append({
                    'Age Group': age_group,
                    'Gender': gender,
                    'Winner': winner['Name'],
                    'Age': winner['Age'],
                    'Club': winner['Club'],
                    'Total Points': winner['Total_Points'],
                    'Events': winner['Events_Count'],
                    'Categories': winner['Categories_Competed']
                })
    
    df_winners = pd.DataFrame(winners)
    output_file = os.path.join(results_folder, 'championship_age_group_winners.csv')
    df_winners.to_csv(output_file, index=False)
    print(f"✓ Saved: {output_file} ({len(df_winners)} age group winners)")


def export_swimmer_narratives(base_folder: str, df_all: pd.DataFrame, event_gender_map: Dict[str, str]) -> None:
    """Create per-swimmer natural-language narratives and write CSV.

    Output: championship_results/championship_swimmer_narratives.csv
    """
    try:
        events = df_all.copy()
        events['Event Number'] = events['Event Number'].astype(str)
        events['Gender'] = events['Event Number'].map(event_gender_map)
        out_rows: List[Dict] = []

        categories = ['Sprint', 'Free', '100 Form', '200 Form', 'IM', 'Distance']

        for swimmer, swimmer_events in events.groupby('Name'):
            swimmer_events = swimmer_events.sort_values('WA Points', ascending=False)
            age = int(swimmer_events['Age'].iloc[0])
            gender = swimmer_events['Gender'].iloc[0]

            # Deduplicate per event, then keep per-category top N
            dedup = swimmer_events.drop_duplicates(subset=['Event Number'], keep='first')
            limit_per_category = 3 if age < 12 else 2
            selected = []
            for cat in categories:
                cat_df = dedup[dedup['Event Category'] == cat]
                if len(cat_df) > 0:
                    selected.append(cat_df.head(limit_per_category))
            included_numbers: List[str] = []
            top8 = None
            if selected:
                all_cand = pd.concat(selected, ignore_index=True)
                top8 = all_cand.nlargest(8, 'WA Points')
                included_numbers = top8['Event Number'].astype(str).tolist()

            total_pts = int(top8['WA Points'].sum()) if top8 is not None and len(top8) else 0
            avg_pts = float(top8['WA Points'].mean()) if top8 is not None and len(top8) else 0.0
            best_pts = int(top8['WA Points'].max()) if top8 is not None and len(top8) else 0

            # Included list by category
            included_by_cat: Dict[str, List[str]] = {}
            if top8 is not None:
                for _, row in top8.iterrows():
                    included_by_cat.setdefault(row['Event Category'], []).append(
                        f"{row['Event Name']} – {int(row['WA Points'])} pts"
                    )

            def _join(items: List[str], max_items: int = 2) -> str:
                items = [i for i in items if i]
                if not items:
                    return ''
                if len(items) > max_items:
                    return ", ".join(items[:max_items]) + f" and {len(items)-max_items} more"
                if len(items) == 1:
                    return items[0]
                return ", ".join(items[:-1]) + f" and {items[-1]}"

            parts = []
            for cat in categories:
                if cat in included_by_cat:
                    parts.append(f"{cat} (" + _join(included_by_cat[cat], 2) + ")")
            included_short = _join(parts, 4) if parts else "no events yet counted"

            out_rows.append({
                'Name': swimmer,
                'Age': age,
                'Gender': gender,
                'Total_Points': total_pts,
                'IncludedShort': included_short,
                'Average_Points': avg_pts,
                'Best_Event_Points': best_pts,
            })

        df_narr = pd.DataFrame(out_rows)
        out_dir = os.path.join(base_folder, 'championship_results')
        os.makedirs(out_dir, exist_ok=True)
        out_path = os.path.join(out_dir, 'championship_swimmer_narratives.csv')
        df_narr.to_csv(out_path, index=False)
        print(f"✓ Saved swimmer narratives: {out_path} ({len(df_narr)} swimmers)")
    except Exception as e:
        print(f"⚠️ Failed to export swimmer narratives: {e}")


def main():
    """Main function to run championship scoreboard calculation."""
    print("=" * 100)
    print("🏆 CLUB CHAMPIONSHIPS SCOREBOARD")
    print("=" * 100)
    
    # Configuration
    # Base championship folder; script will read cleaned_files/ within this
    base_folder = 'WSC_Club_Champs_2024'
    excel_file = 'WSC Club Champs 2024.xlsx'  # optional; ignored if missing
    events_folder = base_folder  # load_all_events will pick cleaned_files/ automatically
    output_folder = base_folder
    
    print("\n📊 Loading event data...")
    
    # Get gender mapping (try Excel, fallback to CSV inference)
    # Always infer gender mapping from cleaned CSVs to avoid Excel dependency
    event_gender_map = get_event_gender_map_from_csvs(base_folder)
    print(f"✓ Mapped {len(event_gender_map)} events to gender (from CSVs)")
    
    # Load all events
    df_all = load_all_events(events_folder)
    print(f"✓ Loaded {len(df_all)} total entries from {df_all['Event Number'].nunique()} events")

    # Export a single unioned file for the dashboard to load efficiently
    export_all_events_union(base_folder, df_all)
    
    # Calculate championship scores
    print("\n🏊 Calculating championship scores...")
    print("Rules:")
    print("  • Top 8 events across at least 5 of 6 categories")
    print("  • Under 12s: max 3 races per category")
    print("  • 12 and over: max 2 races per category")
    
    df_champs = calculate_championship_scores(df_all, event_gender_map)
    print(f"✓ {len(df_champs)} swimmers eligible for championship")
    
    # Create age groups
    df_champs = create_age_groups(df_champs)
    
    # Display scoreboards
    display_scoreboard(df_champs, 'Male', '🏊‍♂️ BOYS CHAMPIONSHIP SCOREBOARD')
    display_scoreboard(df_champs, 'Female', '🏊‍♀️ GIRLS CHAMPIONSHIP SCOREBOARD')
    
    # Export results
    print("\n" + "=" * 100)
    print("📁 EXPORTING RESULTS")
    print("=" * 100)
    export_scoreboard(df_champs, output_folder)

    # Export narratives for dashboard tooltips
    export_swimmer_narratives(base_folder, df_all, event_gender_map)
    
    # Summary statistics
    print("\n" + "=" * 100)
    print("📊 SUMMARY STATISTICS")
    print("=" * 100)
    
    for age_group in ['9-10', '11-12', '13-14', '15-16', '17+']:
        boys_count = len(df_champs[(df_champs['Age Group'] == age_group) & (df_champs['Gender'] == 'Male')])
        girls_count = len(df_champs[(df_champs['Age Group'] == age_group) & (df_champs['Gender'] == 'Female')])
        print(f"{age_group:<8} - Boys: {boys_count:<3} Girls: {girls_count:<3} Total: {boys_count + girls_count}")
    
    print("\n" + "=" * 100)
    print("✅ CHAMPIONSHIP SCOREBOARD COMPLETE!")
    print("=" * 100)


if __name__ == '__main__':
    main()

