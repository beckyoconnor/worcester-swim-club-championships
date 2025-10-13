"""
Swim Event Data Extractor
==========================

A library to extract and clean swimming event data from Excel files, and now
also from MeetManager-style .RES text files. Automatically detects numbered
event sheets (Excel) or parses event blocks (RES) and extracts swimmer
information including name, age, club, time and WA points.

Usage (Excel):
    from swim_event_extractor import SwimEventExtractor
    
    extractor = SwimEventExtractor('WSC Club Champs 2024.xlsx')
    extractor.extract_all_events()

Usage (RES files):
    from swim_event_extractor import SwimEventExtractor

    extractor = SwimEventExtractor(excel_file='RES', output_dir='WSC_Club_Champs_2025', auto_create_folder=False)
    extractor.extract_all_events_from_res('WSC_Club_Champs_2025/raw_files')

Or run from command line:
    python swim_event_extractor.py <excel_file>
    python swim_event_extractor.py --res-dir <folder> [--output-dir <folder>]
"""

import csv
import os
import re
import sys
from datetime import datetime
from typing import List, Dict, Optional, Tuple
try:
    from openpyxl import load_workbook as _oxl_load_workbook
    from openpyxl.worksheet.worksheet import Worksheet as _Worksheet
except Exception:
    _oxl_load_workbook = None
    _Worksheet = object


class SwimEventExtractor:
    """Extract and clean swimming event data from Excel files."""
    
    def __init__(self, excel_file: str, output_dir: str = None, auto_create_folder: bool = True):
        """
        Initialize the extractor.
        
        Args:
            excel_file: Path to the Excel file containing event data
            output_dir: Directory to save CSV files (defaults to auto-generated folder)
            auto_create_folder: If True, creates a folder with competition name, date, and filename
        """
        self.excel_file = excel_file
        self.auto_create_folder = auto_create_folder
        self.workbook = None
        self.competition_info = None
        
        # Determine output directory
        if output_dir:
            self.output_dir = output_dir
        elif auto_create_folder:
            self.output_dir = None  # Will be set after extracting competition info
        else:
            self.output_dir = os.path.dirname(excel_file) or '.'
        
    def load_workbook(self):
        """Load the Excel workbook."""
        if _oxl_load_workbook is None:
            raise RuntimeError("openpyxl is required for Excel mode but is not installed. Use --res-dir for RES parsing.")
        print(f"Loading workbook: {self.excel_file}")
        self.workbook = _oxl_load_workbook(self.excel_file)
    
    def extract_competition_info(self) -> Dict[str, str]:
        """
        Extract competition information from the Excel file.
        Looks for competition name and date in the first sheet or filename.
        
        Returns:
            Dictionary with competition_name, date, and excel_filename
        """
        if not self.workbook:
            self.load_workbook()
        
        competition_name = None
        date_str = None
        
        # Try to get competition name from first few sheets
        # Check first 5 sheets (skip numeric event sheets)
        for sheet_name in self.workbook.sheetnames[:5]:
            if sheet_name.isdigit():
                continue
                
            sheet = self.workbook[sheet_name]
            
            # Look in first few rows for competition name
            for row_idx in range(1, min(10, sheet.max_row + 1)):
                for col_idx in range(1, 4):  # Check first 3 columns
                    cell_value = sheet.cell(row_idx, col_idx).value
                    if cell_value and isinstance(cell_value, str):
                        cell_value = str(cell_value).strip()
                        # Look for competition name patterns
                        if any(word in cell_value.lower() for word in ['championship', 'champs', 'gala', 'meet']):
                            # Clean up the competition name
                            competition_name = cell_value
                            # Try to extract year/date from the name
                            year_match = re.search(r'20\d{2}', competition_name)
                            if year_match:
                                date_str = year_match.group()
                            break
                if competition_name:
                    break
            if competition_name:
                break
        
        # If no competition name found, use a default
        if not competition_name:
            competition_name = "Swimming Competition"
        
        # If no date found in competition name, try to get file modification date
        if not date_str:
            try:
                mod_time = os.path.getmtime(self.excel_file)
                date_str = datetime.fromtimestamp(mod_time).strftime('%Y-%m-%d')
            except:
                date_str = datetime.now().strftime('%Y-%m-%d')
        
        # Get base filename without extension
        excel_filename = os.path.splitext(os.path.basename(self.excel_file))[0]
        
        self.competition_info = {
            'competition_name': competition_name,
            'date': date_str,
            'excel_filename': excel_filename
        }
        
        return self.competition_info
    
    @staticmethod
    def sanitize_folder_name(name: str) -> str:
        """
        Sanitize a string to be used as a folder name.
        
        Args:
            name: The string to sanitize
            
        Returns:
            Sanitized string safe for use as folder name
        """
        # Remove or replace invalid characters
        name = re.sub(r'[<>:"/\\|?*]', '', name)
        # Replace spaces with underscores
        name = re.sub(r'\s+', '_', name)
        # Remove multiple underscores
        name = re.sub(r'_+', '_', name)
        # Trim leading/trailing underscores
        name = name.strip('_')
        # Limit length
        if len(name) > 100:
            name = name[:100]
        return name
    
    def create_output_folder(self) -> str:
        """
        Create the output folder based on competition info.
        
        Returns:
            Path to the created output folder
        """
        if not self.competition_info:
            self.extract_competition_info()
        
        # Build folder name: CompetitionName_Date_ExcelFilename
        folder_parts = []
        
        comp_name = self.competition_info['competition_name']
        date = self.competition_info['date']
        excel_filename = self.competition_info['excel_filename']
        
        # Simplify competition name if it's too long
        if len(comp_name) > 50:
            # Try to extract key words
            words = comp_name.split()
            comp_name = ' '.join(words[:5])  # Take first 5 words
        
        # Sanitize all parts
        comp_name_sanitized = self.sanitize_folder_name(comp_name)
        excel_filename_sanitized = self.sanitize_folder_name(excel_filename)
        
        # Add competition name
        folder_parts.append(comp_name_sanitized)
        
        # Add date if it's not already in the competition name
        if date not in comp_name:
            folder_parts.append(date)
        
        # Add excel filename only if it's different from competition name
        # (to avoid redundancy like "WSC_Club_Champs_2024_2024_WSC_Club_Champs_2024")
        if comp_name_sanitized.lower() != excel_filename_sanitized.lower():
            folder_parts.append(excel_filename_sanitized)
        
        folder_name = '_'.join(folder_parts)
        
        # Create folder in the same directory as the Excel file
        excel_dir = os.path.dirname(os.path.abspath(self.excel_file))
        output_path = os.path.join(excel_dir, folder_name)
        
        # Create the folder if it doesn't exist
        os.makedirs(output_path, exist_ok=True)
        
        return output_path
        
    def get_numbered_sheets(self) -> List[str]:
        """
        Get all sheet names that are numeric (representing events).
        
        Returns:
            List of sheet names that are numbers
        """
        if not self.workbook:
            self.load_workbook()
        
        numbered_sheets = [sheet for sheet in self.workbook.sheetnames if sheet.isdigit()]
        return sorted(numbered_sheets)
    
    def find_columns(self, worksheet: _Worksheet) -> Optional[Dict[str, int]]:
        """
        Find the column indices for Name, Age, Club, Time, and WA Points.
        
        Args:
            worksheet: The worksheet to search
            
        Returns:
            Dictionary with column indices or None if not all columns found
        """
        name_col = None
        age_col = None
        club_col = None
        time_col = None
        wa_pts_col = None
        
        # Search first 10 rows for header row
        for row in worksheet.iter_rows(max_row=10, values_only=True):
            for idx, cell in enumerate(row, start=1):
                if cell:
                    cell_str = str(cell).strip().lower()
                    # Name column detection
                    if not name_col:
                        if cell_str == 'name' or cell_str == 'swimmer name' or cell_str == 'swimmer':
                            name_col = idx
                    # Age column detection (fixed typo from 'aad' to 'age')
                    if not age_col:
                        if cell_str == 'age' or cell_str == 'aad' or cell_str == 'swimmer age':
                            age_col = idx
                    # Club column detection
                    if not club_col:
                        if cell_str == 'club' or cell_str == 'team' or cell_str == 'club name':
                            club_col = idx
                    # Time column detection
                    if not time_col:
                        if cell_str == 'time' or cell_str == 'event time':
                            time_col = idx
                    # WA Points column detection
                    if not wa_pts_col:
                        if ('wa' in cell_str and 'pt' in cell_str) or cell_str == 'wa points' or cell_str == 'points':
                            wa_pts_col = idx
            
            # If we found all required columns, break
            # Note: time_col is optional, but we need the others
            if name_col and age_col and club_col and wa_pts_col:
                break
        
        if not all([name_col, age_col, club_col, wa_pts_col]):
            return None
            
        return {
            'name': name_col,
            'age': age_col,
            'club': club_col,
            'time': time_col,
            'wa_points': wa_pts_col
        }
    
    @staticmethod
    def _extract_clean_event_name(event_name: str, sheet_name: str) -> str:
        """
        Extract clean event name from the full event title.
        Removes "EVENT XXX" prefix but preserves gender information.
        
        Args:
            event_name: Full event name from sheet
            sheet_name: Sheet name (event number)
            
        Returns:
            Clean event name with gender (e.g., "Male 400m Freestyle", "Female 100m IM")
        """
        import re
        
        # Remove "EVENT XXX" prefix (e.g., "EVENT 103 Open/Male 400m Freestyle" -> "Open/Male 400m Freestyle")
        clean_name = re.sub(r'^EVENT\s+\d+\s+', '', event_name, flags=re.IGNORECASE)
        
        # Keep gender markers - don't remove them!
        # The gender information is important for the dashboard
        
        # If nothing left or looks wrong, return the original
        if not clean_name or len(clean_name.strip()) < 3:
            # Try to extract just the race description from original
            # Look for patterns like "400m Freestyle", "100m IM", etc.
            match = re.search(r'(\d+m?\s+[\w\s]+)$', event_name, flags=re.IGNORECASE)
            if match:
                clean_name = match.group(1)
            else:
                clean_name = event_name
        
        return clean_name.strip()
    
    @staticmethod
    def _determine_gender_from_event_name(event_name: str) -> str:
        """
        Determine gender from event name.
        
        Args:
            event_name: Clean event name (e.g., "Male 100m IM", "Female 50m Backstroke")
            
        Returns:
            Gender string ("Male/Open", "Female", or "Unknown")
        """
        event_lower = event_name.lower()
        
        if 'female' in event_lower:
            return 'Female'
        elif 'male' in event_lower or 'open' in event_lower:
            return 'Male/Open'
        else:
            return 'Unknown'
    
    @staticmethod
    def _categorize_event(event_name: str) -> str:
        """
        Categorize event based on championship rules.
        
        Categories:
        - Sprint: 50 free, back, breast and fly
        - Free: 100, 200 and 400 free
        - 100 Form: 100m back, breast and fly
        - 200 Form: 200m back, breast and fly
        - IM: 100, 200, and 400 IM
        - Distance: 800 and 1500
        
        Args:
            event_name: Clean event name
            
        Returns:
            Event category
        """
        event_lower = event_name.lower()
        
        # Check for 50m events (Sprint)
        if '50m' in event_lower or '50 m' in event_lower:
            return 'Sprint'
        
        # Check for Distance events (800m and 1500m)
        if '800m' in event_lower or '1500m' in event_lower:
            return 'Distance'
        
        # Check for IM events
        if 'im' in event_lower or 'medley' in event_lower:
            return 'IM'
        
        # Check for Freestyle events (100m, 200m, 400m - excluding 50m Sprint and 800m/1500m Distance)
        if 'freestyle' in event_lower or 'free' in event_lower:
            return 'Free'
        
        # Check for Form events (Backstroke, Breaststroke, Butterfly)
        # Split into 100 Form and 200 Form
        is_form_stroke = any(stroke in event_lower for stroke in ['backstroke', 'breaststroke', 'butterfly', 'back', 'breast', 'fly'])
        
        if is_form_stroke:
            # Determine if it's 100m or 200m
            if '100m' in event_lower or '100 m' in event_lower:
                return '100 Form'
            elif '200m' in event_lower or '200 m' in event_lower:
                return '200 Form'
            else:
                # Default to form if distance unclear
                return '100 Form'
        
        # Default fallback
        return 'Other'
    
    @staticmethod
    def validate_name(name) -> bool:
        """
        Validate that the name is a valid person's name.
        Ensures no NaN, None, or empty values.
        
        Args:
            name: The name to validate
            
        Returns:
            True if valid, False otherwise
        """
        # Check for None, NaN, or empty values
        if name is None or name == '' or (isinstance(name, float) and name != name):
            return False
        
        if not isinstance(name, str):
            return False
        
        name_str = str(name).strip()
        
        # Must contain at least some letters and be at least 2 characters
        if len(name_str) < 2 or not any(c.isalpha() for c in name_str):
            return False
        
        # Skip header-like rows
        name_lower = name_str.lower()
        if name_lower in ['name', 'place', 'swimmer'] or 'age group' in name_lower:
            return False
            
        return True
    
    @staticmethod
    def validate_age(age) -> Optional[int]:
        """
        Validate and convert age to integer.
        Ensures no NaN, None, or invalid values.
        
        Args:
            age: The age value to validate
            
        Returns:
            Integer age if valid (1-99), None otherwise
        """
        # Check for None, NaN, or empty values
        if age is None or age == '' or (isinstance(age, float) and age != age):
            return None
        
        try:
            age_int = int(age)
            return age_int if 1 <= age_int <= 99 else None
        except (ValueError, TypeError):
            return None
    
    @staticmethod
    def validate_club(club) -> Optional[str]:
        """
        Validate club name.
        Ensures no NaN, None, or empty values.
        
        Args:
            club: The club name to validate
            
        Returns:
            Cleaned club name if valid, None otherwise
        """
        # Check for None, NaN, or empty values
        if club is None or club == '' or (isinstance(club, float) and club != club):
            return None
        
        if not isinstance(club, str):
            return None
        
        club_str = str(club).strip()
        return club_str if len(club_str) >= 2 else None
    
    @staticmethod
    def validate_time(time) -> Optional[str]:
        """
        Validate and clean time value.
        Handles various time formats and DNC/DQ entries.
        
        Args:
            time: The time value to validate
            
        Returns:
            Cleaned time string if valid, None if DNC/DQ or invalid
        """
        # Check for None, NaN, or empty values
        if time is None or time == '' or (isinstance(time, float) and time != time):
            return None
        
        # Convert to string and clean
        time_str = str(time).strip()
        
        # Check for disqualifications or did not compete
        if any(x in time_str.upper() for x in ['DNC', 'DNF', 'DQ', 'NS']):
            return None
        
        # If it's a valid time string (contains numbers and possibly colons/periods)
        if any(c.isdigit() for c in time_str):
            return time_str
        
        return None
    
    @staticmethod
    def validate_wa_points(points) -> Optional[int]:
        """
        Validate and convert WA Points to integer.
        Ensures no NaN, None, or invalid values.
        
        Args:
            points: The WA Points value to validate
            
        Returns:
            Integer points if valid (1-999), None otherwise
        """
        # Check for None, NaN, or empty values
        if points is None or points == '' or (isinstance(points, float) and points != points):
            return None
        
        try:
            points_int = int(points)
            return points_int if 1 <= points_int <= 999 else None
        except (ValueError, TypeError):
            return None
    
    def extract_event_data(self, sheet_name: str) -> Tuple[List[Dict], str]:
        """
        Extract data from a single event sheet.
        
        Args:
            sheet_name: Name of the sheet to extract
            
        Returns:
            Tuple of (list of swimmer data dictionaries, event name)
        """
        worksheet = self.workbook[sheet_name]
        
        # Get event name from first several rows (some events have title in different rows)
        event_name = None
        
        # Check first 10 rows for event name
        for row_idx in range(1, 11):
            cell_value = worksheet.cell(row_idx, 1).value
            if cell_value and 'EVENT' in str(cell_value).upper():
                event_name = str(cell_value).strip()
                break
        
        # If no EVENT found, use first row
        if not event_name:
            event_name = worksheet.cell(1, 1).value
            if event_name:
                event_name = str(event_name).strip()
            else:
                event_name = f"Event {sheet_name}"
        
        # Extract clean event name (remove "EVENT XXX" prefix and gender markers)
        event_name_clean = self._extract_clean_event_name(event_name, sheet_name)
        
        # Categorize the event
        event_category = self._categorize_event(event_name_clean)
        
        # Find column positions
        columns = self.find_columns(worksheet)
        if not columns:
            print(f"  ⚠ Warning: Could not find all required columns in sheet {sheet_name}")
            return [], event_name
        
        # Extract and validate data
        data = []
        seen_swimmers = set()  # Track unique swimmer entries to avoid duplicates
        
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < max(columns.values()):
                continue
            
            # Extract values
            name = row[columns['name'] - 1]
            age = row[columns['age'] - 1]
            club = row[columns['club'] - 1]
            time = row[columns['time'] - 1] if columns['time'] else None
            wa_points = row[columns['wa_points'] - 1]
            
            # Validate all fields - must have NO NaN/None values
            if not self.validate_name(name):
                continue
            
            age_validated = self.validate_age(age)
            if age_validated is None:
                continue
            
            club_validated = self.validate_club(club)
            if club_validated is None:
                continue
            
            # Validate and clean time (optional field)
            time_clean = self.validate_time(time)
            
            wa_points_validated = self.validate_wa_points(wa_points)
            if wa_points_validated is None:
                continue
            
            # Create unique key to detect duplicates within same event
            name_clean = str(name).strip()
            unique_key = (sheet_name, name_clean, age_validated, wa_points_validated)
            
            # Skip if we've already seen this exact entry (duplicate)
            if unique_key in seen_swimmers:
                continue
            
            seen_swimmers.add(unique_key)
            
            # All validations passed and no duplicates - add to data
            # Determine gender from event name
            gender = self._determine_gender_from_event_name(event_name_clean)
            
            data.append({
                'Event Number': sheet_name,
                'Event Name': event_name_clean,
                'Event Category': event_category,
                'Gender': gender,
                'Name': name_clean,
                'Age': age_validated,
                'Club': club_validated,
                'Time': time_clean,
                'WA Points': wa_points_validated
            })
        
        return data, event_name
    
    def save_to_csv(self, data: List[Dict], event_number: str) -> str:
        """
        Save event data to a CSV file.
        
        Args:
            data: List of swimmer data dictionaries
            event_number: The event number (used for filename)
            
        Returns:
            Path to the created CSV file
        """
        # Create cleaned_files subfolder if it doesn't exist
        cleaned_files_dir = os.path.join(self.output_dir, 'cleaned_files')
        os.makedirs(cleaned_files_dir, exist_ok=True)
        
        filename = os.path.join(cleaned_files_dir, f"event_{event_number}.csv")
        
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Event Number', 'Event Name', 'Event Category', 'Gender', 'Name', 'Age', 'Club', 'Time', 'WA Points']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        
        return filename
    
    def extract_all_events(self, verbose: bool = True) -> Dict[str, int]:
        """
        Extract all numbered event sheets and save to CSV files.
        
        Args:
            verbose: Whether to print progress information
            
        Returns:
            Dictionary mapping event numbers to swimmer counts
        """
        if not self.workbook:
            self.load_workbook()
        
        # Create output folder if needed
        if self.output_dir is None and self.auto_create_folder:
            self.output_dir = self.create_output_folder()
            if verbose:
                info = self.competition_info
                print(f"\n📁 Created output folder:")
                print(f"   Competition: {info['competition_name']}")
                print(f"   Date: {info['date']}")
                print(f"   Folder: {os.path.basename(self.output_dir)}")
        
        numbered_sheets = self.get_numbered_sheets()
        
        if verbose:
            print(f"\n🏊 Found {len(numbered_sheets)} numbered event sheets")
            print(f"📂 Output directory: {self.output_dir}/cleaned_files\n")
        
        results = {}
        total_swimmers = 0
        
        for sheet_name in numbered_sheets:
            data, event_name = self.extract_event_data(sheet_name)
            
            if data:
                self.save_to_csv(data, sheet_name)
                results[sheet_name] = len(data)
                total_swimmers += len(data)
                
                if verbose:
                    print(f"✓ event_{sheet_name}.csv: {len(data)} swimmers ({event_name})")
            else:
                results[sheet_name] = 0
                if verbose:
                    print(f"⚠ event_{sheet_name}.csv: 0 swimmers ({event_name})")
        
        if verbose:
            print(f"\n{'='*60}")
            print(f"SUMMARY:")
            print(f"  Total events: {len(numbered_sheets)}")
            print(f"  Total swimmers: {total_swimmers}")
            print(f"  CSV files saved to: {self.output_dir}/cleaned_files")
            print(f"{'='*60}")
        
        return results
    
    # =========================
    # RES (.RES) parsing support
    # =========================

    @staticmethod
    def _map_club_code(club_code: str) -> str:
        """
        Map common club codes found in RES files to full club names.
        Defaults to returning the original string if unknown.
        """
        code = (club_code or '').strip().upper()
        mapping = {
            'WORM': 'Worcester',
        }
        return mapping.get(code, club_code)

    def _parse_res_file(self, res_path: str) -> Tuple[str, str, str, str, List[Dict]]:
        """
        Parse a single .RES file into standardized event rows.

        Returns:
            (event_number, event_name_clean, gender, event_category, data_rows)
        """
        event_number: Optional[str] = None
        event_full_name: Optional[str] = None
        data: List[Dict] = []

        try:
            with open(res_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = [ln.rstrip('\n') for ln in f]
        except Exception as e:
            print(f"⚠ Failed to read {res_path}: {e}")
            return '', '', 'Unknown', 'Other', []

        # Find event header like: "EVENT 302 Open/Male 400m Freestyle"
        header_re = re.compile(r'^\s*EVENT\s+(\d{3})\s+(.+?)\s*$', flags=re.IGNORECASE)
        for ln in lines:
            m = header_re.match(ln)
            if m:
                event_number = m.group(1)
                event_full_name = m.group(2).strip()
                break

        if not event_number:
            # If filename is like CC25E302.RES, fall back to extracting digits
            base = os.path.basename(res_path)
            m = re.search(r'E(\d{3})', base, flags=re.IGNORECASE)
            if m:
                event_number = m.group(1)
            else:
                event_number = ''
        if not event_full_name:
            # Try to derive a minimal name from file if header missing
            event_full_name = f"Event {event_number}" if event_number else "Event"

        event_name_clean = self._extract_clean_event_name(f"EVENT {event_number} {event_full_name}", event_number or '')
        gender = self._determine_gender_from_event_name(event_name_clean)
        event_category = self._categorize_event(event_name_clean)

        # Parse result rows. Typical line example (tabs/separated spacing):
        # 1.\tLucy PIPER\t12\tWORM\t 6:04.09\t\t\t359
        # Skip header lines containing 'Place' or 'Age Group'
        place_line_re = re.compile(r'^\s*\d+\.', re.ASCII)
        dnc_markers = ('DNC', 'DNF', 'DQ', 'NS')

        for raw in lines:
            line = raw.strip()
            if not line:
                continue
            upper = line.upper()
            if 'AGE GROUP' in upper or 'PLACE' in upper:
                continue
            if not place_line_re.match(line):
                continue
            # Skip non-scoring rows only if markers appear as standalone tokens
            # (avoid false positives like 'JENKINS' containing 'NS')
            import re as _re
            if _re.search(r'(?:^|\s)(?:DNC|DNF|DQ|NS)(?:\s|$)', upper):
                continue

            # Extract columns via regex aimed at: place, name, age, club, time, wa
            # Name can include spaces and hyphens; club typically 3-4 uppercase letters
            m = re.match(r'^\s*\d+\.\s+(.+?)\s+(\d{1,2})\s+([A-Za-z]{3,4})\s+([0-9:\.]+)\s+(\d+)\s*$', line)
            if not m:
                # Try a more permissive split on tabs / multi-space, then pick last numeric as WA
                parts = [p for p in re.split(r'\t+|\s{2,}', line) if p]
                # Expect at least: place, name, age, club, time, wa
                if len(parts) < 6:
                    continue
                place = parts[0]
                name = parts[1]
                age = parts[2]
                club = parts[3]
                # WA points tends to be last numeric token
                wa_match = None
                for p in reversed(parts):
                    if re.fullmatch(r'\d+', p):
                        wa_match = p
                        break
                if not wa_match:
                    continue
                time_token = None
                # time likely the token just before last numeric
                try:
                    wa_index = parts.index(wa_match)
                    if wa_index - 1 >= 0:
                        time_token = parts[wa_index - 1]
                except ValueError:
                    pass
                if not time_token:
                    continue
            else:
                name = m.group(1).strip()
                age = m.group(2).strip()
                club = m.group(3).strip()
                time_token = m.group(4).strip()
                wa_match = m.group(5).strip()

            # Validate using existing validators
            if not self.validate_name(name):
                continue
            age_valid = self.validate_age(age)
            if age_valid is None:
                continue
            club_clean = self._map_club_code(club)
            if self.validate_club(club_clean) is None:
                continue
            time_clean = self.validate_time(time_token)
            wa_valid = self.validate_wa_points(wa_match)
            if wa_valid is None:
                continue

            data.append({
                'Event Number': event_number,
                'Event Name': event_name_clean,
                'Event Category': event_category,
                'Gender': gender,
                'Name': str(name).strip(),
                'Age': age_valid,
                'Club': club_clean,
                'Time': time_clean,
                'WA Points': wa_valid,
            })

        return event_number or '', event_name_clean, gender, event_category, data

    def extract_all_events_from_res(self, res_dir: str, verbose: bool = True) -> Dict[str, int]:
        """
        Parse all .RES files in a folder and write standardized CSVs to cleaned_files/.
        """
        if not os.path.isdir(res_dir):
            raise FileNotFoundError(f"RES folder not found: {res_dir}")

        # Ensure output directory configured
        if self.output_dir is None:
            # Default to parent of res_dir
            self.output_dir = os.path.abspath(os.path.join(res_dir, os.pardir))

        cleaned_dir = os.path.join(self.output_dir, 'cleaned_files')
        os.makedirs(cleaned_dir, exist_ok=True)

        results: Dict[str, int] = {}
        total_swimmers = 0

        files = sorted([f for f in os.listdir(res_dir) if f.upper().endswith('.RES')])
        if verbose:
            print(f"\n🏊 Found {len(files)} RES files")
            print(f"📂 Output directory: {self.output_dir}/cleaned_files\n")

        for fname in files:
            path = os.path.join(res_dir, fname)
            event_number, event_name, _, _, rows = self._parse_res_file(path)
            if not event_number:
                if verbose:
                    print(f"⚠ Skipping {fname}: could not determine event number")
                continue
            if rows:
                self.save_to_csv(rows, event_number)
                results[event_number] = len(rows)
                total_swimmers += len(rows)
                if verbose:
                    print(f"✓ event_{event_number}.csv: {len(rows)} swimmers ({event_name})")
            else:
                results[event_number] = 0
                if verbose:
                    print(f"⚠ event_{event_number}.csv: 0 swimmers ({event_name})")

        if verbose:
            print(f"\n{'='*60}")
            print("SUMMARY:")
            print(f"  Total events: {len(results)}")
            print(f"  Total swimmers: {total_swimmers}")
            print(f"  CSV files saved to: {self.output_dir}/cleaned_files")
            print(f"{'='*60}")

        return results

    def create_combined_csv(self, output_filename: str = "all_events_combined.csv") -> str:
        """
        Create a single CSV file containing all events.
        
        Args:
            output_filename: Name of the combined CSV file
            
        Returns:
            Path to the created CSV file
        """
        if not self.workbook:
            self.load_workbook()
        
        numbered_sheets = self.get_numbered_sheets()
        all_data = []
        
        print(f"\nCreating combined CSV with all events...")
        
        for sheet_name in numbered_sheets:
            data, _ = self.extract_event_data(sheet_name)
            all_data.extend(data)
        
        output_path = os.path.join(self.output_dir, output_filename)
        
        with open(output_path, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Event Number', 'Name', 'Age', 'Club', 'WA Points']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(all_data)
        
        print(f"✓ Combined CSV created: {output_path}")
        print(f"  Total rows: {len(all_data)}")
        
        return output_path


def main():
    """Command-line interface."""
    # Simple CLI parsing to support both Excel and RES modes
    args = sys.argv[1:]
    if not args:
        print("Usage:")
        print("  python swim_event_extractor.py <excel_file> [output_directory]")
        print("  python swim_event_extractor.py --res-dir <folder> [--output-dir <folder>]")
        sys.exit(1)

    if '--res-dir' in args:
        try:
            res_idx = args.index('--res-dir')
            res_dir = args[res_idx + 1]
        except Exception:
            print("Error: --res-dir requires a folder path")
            sys.exit(1)

        output_dir = None
        if '--output-dir' in args:
            try:
                out_idx = args.index('--output-dir')
                output_dir = args[out_idx + 1]
            except Exception:
                print("Error: --output-dir requires a folder path")
                sys.exit(1)

        extractor = SwimEventExtractor(excel_file='RES', output_dir=output_dir, auto_create_folder=False)
        extractor.extract_all_events_from_res(res_dir)
        print("\n✓ RES processing complete!")
        return

    # Default Excel mode
    excel_file = args[0]
    output_dir = args[1] if len(args) > 1 else None

    if not os.path.exists(excel_file):
        print(f"Error: File '{excel_file}' not found")
        sys.exit(1)

    extractor = SwimEventExtractor(excel_file, output_dir)
    extractor.extract_all_events()

    # Ask if user wants combined CSV
    print("\n" + "="*60)
    response = input("Create a combined CSV with all events? (y/n): ").strip().lower()
    if response == 'y':
        extractor.create_combined_csv()

    print("\n✓ Processing complete!")


if __name__ == "__main__":
    main()

